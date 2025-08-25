from django.shortcuts import render
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework import viewsets, status
from rest_framework.parsers import MultiPartParser, FormParser
from django.db import transaction
from django.http import HttpResponse
from .models import Department, Domain, Question, Assessment, Response as AssessmentResponse, Evidence
from .serializers import (
	DepartmentSerializer, DomainSerializer, QuestionSerializer,
	AssessmentSerializer, ResponseSerializer, EvidenceSerializer
)
from .permissions import IsCoordinatorOrAdmin, ReadOnly, DepartmentScopedQuerysetMixin
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# Create your views here.


@api_view(["GET"])
@permission_classes([AllowAny])
def health(request):
	return Response({"status": "ok"})


class DomainViewSet(viewsets.ModelViewSet):
	queryset = Domain.objects.all()
	serializer_class = DomainSerializer
	permission_classes = [IsAuthenticated]


class QuestionViewSet(viewsets.ModelViewSet):
	queryset = Question.objects.select_related('domain').all()
	serializer_class = QuestionSerializer
	permission_classes = [IsAuthenticated]


class AssessmentViewSet(DepartmentScopedQuerysetMixin, viewsets.ModelViewSet):
	queryset = Assessment.objects.select_related('department', 'owner').all()
	serializer_class = AssessmentSerializer
	permission_classes = [IsAuthenticated]


class ResponseViewSet(DepartmentScopedQuerysetMixin, viewsets.ModelViewSet):
	queryset = AssessmentResponse.objects.select_related('assessment', 'question').all()
	serializer_class = ResponseSerializer
	permission_classes = [IsAuthenticated]

	def perform_create(self, serializer):
		with transaction.atomic():
			instance = serializer.save(created_by=self.request.user, updated_by=self.request.user)
			self._validate_rating_range(instance)

	def perform_update(self, serializer):
		with transaction.atomic():
			instance = serializer.save(updated_by=self.request.user)
			self._validate_rating_range(instance)

	def _validate_rating_range(self, instance: AssessmentResponse):
		if instance.rating is not None and not (0 <= instance.rating <= 5):
			raise ValueError("Rating must be between 0 and 5")


class EvidenceViewSet(DepartmentScopedQuerysetMixin, viewsets.ModelViewSet):
	queryset = Evidence.objects.select_related('response').all()
	serializer_class = EvidenceSerializer
	permission_classes = [IsAuthenticated]
	parser_classes = [MultiPartParser, FormParser]

	def perform_create(self, serializer):
		file_obj = self.request.FILES.get('file')
		if not file_obj:
			raise ValueError("File is required")
		if file_obj.size > 10 * 1024 * 1024:
			raise ValueError("File exceeds 10 MB limit")
		content_type = file_obj.content_type
		instance = serializer.save(uploaded_by=self.request.user, content_type=content_type)
		# increment version per response
		latest = Evidence.objects.filter(response=instance.response).order_by('-version').first()
		instance.version = 1 if latest is None else latest.version + 1
		instance.save(update_fields=['version'])


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_assessments_csv(request):
	response = HttpResponse(content_type='text/csv')
	response["Content-Disposition"] = 'attachment; filename="assessments.csv"'
	writer = csv.writer(response)
	writer.writerow(["Title", "Department", "Status", "Overall Score"]) 
	for a in AssessmentViewSet().get_queryset():
		serializer = AssessmentSerializer(a, context={"request": request})
		writer.writerow([a.title, a.department.name, a.status, serializer.data.get("overall_score")])
	return response


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_assessments_pdf(request):
	# Minimal stub PDF using plain text response; replace with real PDF lib later
	content = "BCM-MAP Assessments\n\n"
	for a in AssessmentViewSet().get_queryset():
		serializer = AssessmentSerializer(a, context={"request": request})
		content += f"- {a.title} | {a.department.name} | {a.status} | score: {serializer.data.get('overall_score')}\n"
	return HttpResponse(content, content_type='application/pdf')


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_assessments_excel(request):
	wb = Workbook()
	ws = wb.active
	ws.title = 'Assessments'
	ws.append(["Title", "Department", "Status", "Overall Score"])
	for a in AssessmentViewSet().get_queryset():
		serializer = AssessmentSerializer(a, context={"request": request})
		ws.append([a.title, a.department.name, a.status, serializer.data.get('overall_score')])
	for col in range(1, 5):
		ws.column_dimensions[get_column_letter(col)].width = 24
	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename="assessments.xlsx"'
	wb.save(response)
	return response
